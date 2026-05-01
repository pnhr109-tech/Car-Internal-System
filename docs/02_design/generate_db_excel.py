"""
docs/02_design/generate_db_excel.py
db-schema.md を解析して db-schema.xlsx を生成するスクリプト。

実行:
    python docs/02_design/generate_db_excel.py
"""
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

# ── カラーパレット ────────────────────────────────────────────
C_HEADER_BG   = "1F4E79"   # 濃紺 (見出し行背景)
C_HEADER_FG   = "FFFFFF"   # 白  (見出し行文字)
C_SECTION_BG  = "D6E4F0"   # 薄青 (テーブル名行)
C_WARN_BG     = "FFF2CC"   # 薄黄 (将来実装テーブル)
C_ALT_BG      = "F5F9FF"   # 超薄青 (偶数行)
C_BORDER      = "AAAAAA"   # 薄グレー (罫線)

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def bold_font(size=10, color="000000") -> Font:
    return Font(bold=True, size=size, color=color)


def normal_font(size=10, color="000000") -> Font:
    return Font(size=size, color=color)


def apply_header_row(ws, row_idx: int, values: list[str]) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = bold_font(10, C_HEADER_FG)
        cell.fill = fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def apply_data_row(ws, row_idx: int, values: list[str], alt: bool = False) -> None:
    bg = C_ALT_BG if alt else "FFFFFF"
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = normal_font()
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER


def set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════
# Markdown パーサ
# ══════════════════════════════════════════════════════════════

def parse_md_tables(md_text: str) -> dict:
    """
    markdown からテーブル情報を抽出する。
    返り値:
        {
          "overview":    [{"#":…, "テーブル名":…, …}, …],
          "tables": [
              {"name": "stores", "section": "accounts", "description": "…",
               "future": False, "columns": [{"カラム":…, "型":…, …}, …],
               "indexes": [...], "extras": [...] },
              …
          ],
          "biz_flow": [{"フェーズ":…, …}, …],
        }
    """
    result = {"overview": [], "tables": [], "biz_flow": []}

    # ── テーブル一覧 ─────────────────────────────────────────
    sec_re = re.compile(r'## 1\. テーブル一覧(.*?)## 2\.', re.DOTALL)
    m = sec_re.search(md_text)
    if m:
        for row in parse_table_rows(m.group(1)):
            result["overview"].append(row)

    # ── 業務フロー ────────────────────────────────────────────
    biz_re = re.compile(r'## 5\. 業務フロー別テーブル対応(.*?)$', re.DOTALL)
    m = biz_re.search(md_text)
    if m:
        for row in parse_table_rows(m.group(1)):
            result["biz_flow"].append(row)

    # ── 各テーブル詳細 ───────────────────────────────────────
    # ### 2.1 `stores` — 店舗マスタ  のような見出しを検出
    tbl_header_re = re.compile(
        r'### (\d+\.\d+) `([^`]+)` — (.+?)\n(.*?)(?=\n### |\Z)',
        re.DOTALL
    )
    for m in tbl_header_re.finditer(md_text):
        _num, tbl_name, desc, body = m.group(1), m.group(2), m.group(3).strip(), m.group(4)

        # accounts か leads か判定
        section_num = int(_num.split('.')[0])
        section = "accounts" if section_num == 2 else "leads"

        future = "将来実装" in body or "将来実装" in desc

        # カラムテーブル（複数あり得る: 基本情報・契約オプション等）
        all_cols = []
        # サブセクション見出し（#### ... ）でグループ分け
        subsec_re = re.compile(r'#### (.+?)\n(.*?)(?=\n#### |\Z)', re.DOTALL)
        subsecs = list(subsec_re.finditer(body))

        if subsecs:
            for sm in subsecs:
                sub_title = sm.group(1).strip()
                sub_body = sm.group(2)
                cols = parse_table_rows(sub_body)
                for c in cols:
                    c["グループ"] = sub_title
                all_cols.extend(cols)
        else:
            cols = parse_table_rows(body)
            for c in cols:
                c["グループ"] = ""
            all_cols.extend(cols)

        # インデックステーブル
        indexes = []
        idx_re = re.compile(r'インデックス.*?\n(\|.*?\n)+', re.DOTALL)
        im = idx_re.search(body)
        if im:
            indexes = parse_table_rows(im.group(0))

        # 制約・ステータス遷移などの補足テキスト
        extras = []
        constraint_re = re.compile(r'\*\*制約\*\*\n+(.*?)(?=\n\n|\Z)', re.DOTALL)
        cm = constraint_re.search(body)
        if cm:
            extras.append(("制約", cm.group(1).strip()))
        status_re = re.compile(r'\*\*ステータス遷移\*\*\n+```\n(.*?)```', re.DOTALL)
        sm2 = status_re.search(body)
        if sm2:
            extras.append(("ステータス遷移", sm2.group(1).strip()))

        result["tables"].append({
            "name": tbl_name,
            "section": section,
            "number": _num,
            "description": desc,
            "future": future,
            "columns": all_cols,
            "indexes": indexes,
            "extras": extras,
        })

    return result


def parse_table_rows(text: str) -> list[dict]:
    """Markdown テーブルを [{col: val, ...}] に変換する。"""
    rows = []
    headers = []
    for line in text.splitlines():
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.strip("|").split("|")]
        if not headers:
            headers = cells
            continue
        if set(c.strip("-: ") for c in cells) == set():  # 区切り行
            continue
        if all(re.match(r'^[-: ]+$', c) for c in cells):
            continue
        if len(cells) == len(headers):
            rows.append(dict(zip(headers, cells)))
    return rows


# ══════════════════════════════════════════════════════════════
# Excel 書き込み
# ══════════════════════════════════════════════════════════════

def write_overview_sheet(wb: openpyxl.Workbook, data: list[dict]) -> None:
    ws = wb.create_sheet("テーブル一覧")
    ws.freeze_panes = "A2"
    headers = ["#", "テーブル名", "モデル", "アプリ", "概要"]
    apply_header_row(ws, 1, headers)
    set_col_widths(ws, [5, 28, 28, 12, 30])
    for i, row in enumerate(data, 2):
        apply_data_row(ws, i, [row.get(h, "") for h in headers], alt=(i % 2 == 0))
    ws.row_dimensions[1].height = 20


def write_table_sheet(wb: openpyxl.Workbook, tbl: dict) -> None:
    sheet_name = tbl["name"][:31]  # Excel シート名 31 文字制限
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes = "A3"

    row = 1

    # タイトル行
    title = f"{tbl['number']}  {tbl['name']}  —  {tbl['description']}"
    if tbl["future"]:
        title += "  ⚠️ 将来実装"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = bold_font(12, "000000")
    cell.fill = fill(C_WARN_BG if tbl["future"] else C_SECTION_BG)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = BORDER
    ws.row_dimensions[row].height = 22
    row += 1

    # カラムヘッダー
    has_group = any(c.get("グループ") for c in tbl["columns"])
    if has_group:
        col_headers = ["グループ", "カラム", "型", "NULL", "デフォルト", "説明"]
        col_widths = [18, 30, 28, 8, 14, 45]
    else:
        col_headers = ["カラム", "型", "NULL", "デフォルト", "説明"]
        col_widths = [30, 28, 8, 14, 45]

    apply_header_row(ws, row, col_headers)
    ws.row_dimensions[row].height = 18
    row += 1

    # カラムデータ
    prev_group = None
    for i, col in enumerate(tbl["columns"]):
        if has_group:
            group = col.get("グループ", "")
            if group and group != prev_group:
                # グループ区切り行
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                gc = ws.cell(row=row, column=1, value=f"◆ {group}")
                gc.font = bold_font(10, "1F4E79")
                gc.fill = fill("EBF3FB")
                gc.alignment = Alignment(horizontal="left", vertical="center")
                gc.border = BORDER
                ws.row_dimensions[row].height = 16
                row += 1
                prev_group = group
            values = [group, col.get("カラム", ""), col.get("型", ""),
                      col.get("NULL", ""), col.get("デフォルト", ""), col.get("説明", "")]
        else:
            values = [col.get("カラム", ""), col.get("型", ""),
                      col.get("NULL", ""), col.get("デフォルト", ""), col.get("説明", "")]

        apply_data_row(ws, row, values, alt=(i % 2 == 0))
        ws.row_dimensions[row].height = 15
        row += 1

    # インデックス
    if tbl["indexes"]:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        hc = ws.cell(row=row, column=1, value="インデックス")
        hc.font = bold_font(11)
        hc.fill = fill(C_SECTION_BG)
        hc.alignment = Alignment(horizontal="left", vertical="center")
        hc.border = BORDER
        ws.row_dimensions[row].height = 18
        row += 1

        idx_headers = list(tbl["indexes"][0].keys()) if tbl["indexes"] else ["名前", "カラム"]
        apply_header_row(ws, row, idx_headers)
        row += 1
        for ii, idx in enumerate(tbl["indexes"]):
            apply_data_row(ws, row, list(idx.values()), alt=(ii % 2 == 0))
            row += 1

    # 補足情報
    for label, content in tbl["extras"]:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        hc = ws.cell(row=row, column=1, value=label)
        hc.font = bold_font(11)
        hc.fill = fill(C_SECTION_BG)
        hc.border = BORDER
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cc = ws.cell(row=row, column=1, value=content)
        cc.font = normal_font()
        cc.alignment = Alignment(wrap_text=True)
        cc.border = BORDER
        ws.row_dimensions[row].height = max(15, content.count("\n") * 15 + 15)
        row += 1

    set_col_widths(ws, col_widths)


def write_bizflow_sheet(wb: openpyxl.Workbook, data: list[dict]) -> None:
    ws = wb.create_sheet("業務フロー別テーブル対応")
    ws.freeze_panes = "A2"
    headers = ["フェーズ", "主テーブル", "補助テーブル"]
    apply_header_row(ws, 1, headers)
    set_col_widths(ws, [28, 35, 55])
    for i, row in enumerate(data, 2):
        apply_data_row(ws, i, [row.get(h, "") for h in headers], alt=(i % 2 == 0))
        ws.row_dimensions[i].height = 18
    ws.row_dimensions[1].height = 20


def write_index_sheet(wb: openpyxl.Workbook, tables: list[dict]) -> None:
    """先頭シート: ナビゲーション用インデックス"""
    ws = wb.create_sheet("INDEX", 0)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 30

    # タイトル
    ws.merge_cells("A1:E1")
    tc = ws.cell(row=1, column=1, value="DB設計書  —  Car Internal System")
    tc.font = bold_font(14, C_HEADER_FG)
    tc.fill = fill(C_HEADER_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.cell(row=2, column=1, value="")

    apply_header_row(ws, 3, ["#", "テーブル名", "モデル名(Django)", "アプリ", "概要"])
    ws.row_dimensions[3].height = 20

    for i, t in enumerate(tables, 1):
        row = i + 3
        # テーブル名をハイパーリンクでシートへジャンプ
        name_cell = ws.cell(row=row, column=2, value=t["name"])
        name_cell.hyperlink = f"#{t['name'][:31]}!A1"
        name_cell.font = Font(size=10, color="0563C1", underline="single")
        name_cell.fill = fill(C_WARN_BG if t["future"] else ("F5F9FF" if i % 2 == 0 else "FFFFFF"))
        name_cell.border = BORDER
        name_cell.alignment = Alignment(horizontal="left", vertical="center")

        other_cols = [str(i), "", t["section"], t["description"]]
        for ci, val in enumerate([str(i), "", t["section"], t["description"]], 1):
            if ci == 2:
                continue  # name already set
            c = ws.cell(row=row, column=ci if ci < 2 else ci + 1, value=val)
            c.font = normal_font()
            c.fill = fill(C_WARN_BG if t["future"] else ("F5F9FF" if i % 2 == 0 else "FFFFFF"))
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 15


# ══════════════════════════════════════════════════════════════
# エントリポイント
# ══════════════════════════════════════════════════════════════

def main():
    base = Path(__file__).parent
    md_path  = base / "db-schema.md"
    out_path = base / "db-schema.xlsx"

    print(f"読み込み: {md_path}")
    md_text = md_path.read_text(encoding="utf-8")

    parsed = parse_md_tables(md_text)
    print(f"  テーブル一覧: {len(parsed['overview'])} 件")
    print(f"  テーブル詳細: {len(parsed['tables'])} 件")
    print(f"  業務フロー:   {len(parsed['biz_flow'])} 件")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    write_index_sheet(wb, parsed["tables"])
    write_overview_sheet(wb, parsed["overview"])
    for tbl in parsed["tables"]:
        write_table_sheet(wb, tbl)
    write_bizflow_sheet(wb, parsed["biz_flow"])

    wb.save(out_path)
    print(f"✅ 生成完了: {out_path}")


if __name__ == "__main__":
    main()
